"""
AB Intelligence — Bulgarian Business Lead Finder
=================================================
Finds Bulgarian businesses that SHOULD have a website but don't.

Source: OpenStreetMap (Overpass API) — free, open, no API key needed.
        Contains thousands of real Bulgarian businesses with addresses,
        phones, and website info.

How it works:
  1. Queries OpenStreetMap for businesses in Bulgaria by type
  2. Filters out those that already have a website listed
  3. Verifies the ones that do have a listed site still work
  4. Saves businesses WITHOUT a working website → your prospects

Requirements:
    pip3 install requests openpyxl tqdm

Output (saved in this folder):
    leads.xlsx        — businesses with NO website (your prospects)
    all_results.xlsx  — everything found

Usage:
    python3 find_leads.py
"""

import requests
import time
import json
import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# ── Settings ──────────────────────────────────────────────────────────────────

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
PROGRESS_FILE = os.path.join(SCRIPT_DIR, "progress.json")
OUTPUT_LEADS  = os.path.join(SCRIPT_DIR, "leads.xlsx")
OUTPUT_ALL    = os.path.join(SCRIPT_DIR, "all_results.xlsx")

OVERPASS_URL  = "https://overpass-api.de/api/interpreter"
DELAY         = 2.0   # seconds between website checks

HEADERS = {
    "User-Agent": "AB-Intelligence-LeadFinder/1.0 (business research tool)",
    "Accept-Language": "bg,en;q=0.9",
}

# ── Business types to query ───────────────────────────────────────────────────
# Each entry: (Category label, OSM tag key, OSM tag value)
# Full list: https://wiki.openstreetmap.org/wiki/Map_features

BUSINESS_TYPES = [
    # Food & drink
    ("Ресторанти",           "amenity",  "restaurant"),
    ("Кафенета",             "amenity",  "cafe"),
    ("Барове",               "amenity",  "bar"),
    ("Хотели",               "tourism",  "hotel"),
    ("Гостилници",           "tourism",  "guest_house"),

    # Health & beauty
    ("Фризьорски салони",    "shop",     "hairdresser"),
    ("Козметични салони",    "shop",     "beauty"),
    ("Аптеки",               "amenity",  "pharmacy"),
    ("Зъболекари",           "amenity",  "dentist"),
    ("Клиники",              "amenity",  "clinic"),
    ("Ветеринарни клиники",  "amenity",  "veterinary"),

    # Retail
    ("Мебели",               "shop",     "furniture"),
    ("Цветарници",           "shop",     "florist"),
    ("Оптики",               "shop",     "optician"),
    ("Бижутерии",            "shop",     "jewelry"),
    ("Спортни стоки",        "shop",     "sports"),
    ("Книжарници",           "shop",     "books"),
    ("Електроника",          "shop",     "electronics"),
    ("Строителни материали", "shop",     "doityourself"),
    ("Супермаркети",         "shop",     "supermarket"),

    # Auto
    ("Автосервизи",          "shop",     "car_repair"),
    ("Автомобили",           "shop",     "car"),
    ("Гуми",                 "shop",     "tyres"),

    # Professional & services
    ("Адвокати",             "office",   "lawyer"),
    ("Счетоводство",         "office",   "accountant"),
    ("Застраховане",         "office",   "insurance"),
    ("Недвижими имоти",      "office",   "estate_agent"),
    ("Туристически агенции", "shop",     "travel_agency"),
    ("Печатници",            "shop",     "copyshop"),
    ("Почистване",           "shop",     "dry_cleaning"),

    # Construction
    ("Строителни фирми",     "shop",     "construction"),

    # Education
    ("Езикови школи",        "amenity",  "language_school"),
    ("Фитнес зали",          "leisure",  "fitness_centre"),
    ("Спортни клубове",      "leisure",  "sports_centre"),
]


# ── Overpass query ────────────────────────────────────────────────────────────

def query_overpass(tag_key: str, tag_value: str) -> list:
    """
    Query all nodes/ways in Bulgaria with the given OSM tag.
    Returns list of dicts with name, phone, website, address fields.
    """
    query = f"""
    [out:json][timeout:60];
    area["ISO3166-1"="BG"]->.bg;
    (
      node["{tag_key}"="{tag_value}"](area.bg);
      way["{tag_key}"="{tag_value}"](area.bg);
    );
    out body center;
    """
    try:
        r = requests.post(OVERPASS_URL, data={"data": query},
                          headers=HEADERS, timeout=90)
        r.raise_for_status()
        data = r.json()

        results = []
        for el in data.get("elements", []):
            tags = el.get("tags", {})
            name = tags.get("name") or tags.get("name:bg") or tags.get("name:en")
            if not name:
                continue

            # Build address
            addr_parts = [
                tags.get("addr:city", ""),
                tags.get("addr:street", ""),
                tags.get("addr:housenumber", ""),
            ]
            address = ", ".join(p for p in addr_parts if p)

            results.append({
                "name":    name.strip(),
                "phone":   tags.get("phone", tags.get("contact:phone", "")),
                "email":   tags.get("email", tags.get("contact:email", "")),
                "website": tags.get("website", tags.get("contact:website",
                           tags.get("url", ""))),
                "address": address,
                "city":    tags.get("addr:city", ""),
                "osm_id":  str(el.get("id", "")),
            })
        return results

    except Exception as e:
        print(f"    Overpass error: {e}")
        return []


# ── Website check ─────────────────────────────────────────────────────────────

def check_url(url: str) -> bool:
    """Return True if the URL responds successfully."""
    if not url:
        return False
    if not url.startswith("http"):
        url = "https://" + url
    try:
        r = requests.head(url, headers=HEADERS, timeout=7, allow_redirects=True)
        return r.status_code < 400
    except:
        try:
            r = requests.get(url, headers=HEADERS, timeout=7, stream=True)
            r.close()
            return r.status_code < 400
        except:
            return False


# ── Progress ──────────────────────────────────────────────────────────────────

def load_progress() -> dict:
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"done": [], "results": []}


def save_progress(p: dict):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(p, f, ensure_ascii=False, indent=2)


# ── Excel output ──────────────────────────────────────────────────────────────

def make_style():
    thin = Side(style="thin", color="CCCCCC")
    return {
        "border":      Border(left=thin, right=thin, top=thin, bottom=thin),
        "header_font": Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "header_fill": PatternFill("solid", fgColor="1A3C5E"),
        "row_font":    Font(name="Arial", size=9),
        "alt_fill":    PatternFill("solid", fgColor="EAF2FB"),
        "no_fill":     PatternFill("solid", fgColor="FFFFFF"),
        "green_fill":  PatternFill("solid", fgColor="D5F0D5"),
        "red_fill":    PatternFill("solid", fgColor="FDE8E8"),
        "center":      Alignment(horizontal="center", vertical="center"),
        "left":        Alignment(horizontal="left",   vertical="center"),
        "url_font":    Font(name="Arial", size=9, color="0563C1", underline="single"),
    }


def write_excel(filepath: str, data: list, sheet_title: str):
    st = make_style()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    headers    = ["#", "Категория", "Наименование", "Телефон", "Имейл", "Град", "Адрес", "Уебсайт?"]
    col_widths = [4,   20,          40,              18,        28,      16,     35,       10]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = st["header_font"]
        c.fill = st["header_fill"]
        c.alignment = st["center"]
        c.border = st["border"]
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for i, row in enumerate(data, 1):
        r = i + 1
        base = st["alt_fill"] if i % 2 == 0 else st["no_fill"]
        has = row.get("has_website", False)
        values = [
            i,
            row.get("category", ""),
            row.get("name", ""),
            row.get("phone", ""),
            row.get("email", ""),
            row.get("city", ""),
            row.get("address", ""),
            "ДА" if has else "НЕ",
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = st["border"]
            if col == 8:
                c.font  = st["row_font"]
                c.fill  = st["green_fill"] if val == "ДА" else st["red_fill"]
            else:
                c.font = st["row_font"]
                c.fill = base
            c.alignment = st["center"] if col == 1 else st["left"]
        ws.row_dimensions[r].height = 15

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    wb.save(filepath)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  AB Intelligence — Bulgarian Business Lead Finder")
    print("=" * 60)
    print(f"  Business categories  : {len(BUSINESS_TYPES)}")
    print(f"  Data source          : OpenStreetMap (Overpass API)")
    print(f"  Output               : {SCRIPT_DIR}")
    print("=" * 60 + "\n")

    progress = load_progress()
    done     = set(progress["done"])
    results  = progress["results"]

    todo = [(label, k, v) for label, k, v in BUSINESS_TYPES
            if f"{k}={v}" not in done]

    already_leads = sum(1 for r in results if not r.get("has_website"))
    print(f"Done: {len(done)}/{len(BUSINESS_TYPES)} categories  |  Leads so far: {already_leads}\n")

    for label, tag_key, tag_value in tqdm(todo, desc="Fetching categories", unit="cat"):
        tqdm.write(f"\n→ {label} [{tag_key}={tag_value}]")

        businesses = query_overpass(tag_key, tag_value)
        tqdm.write(f"  {len(businesses)} businesses found in OpenStreetMap")

        if not businesses:
            done.add(f"{tag_key}={tag_value}")
            progress["done"] = list(done)
            save_progress(progress)
            continue

        # Deduplicate by OSM id
        existing_ids = {r["osm_id"] for r in results if "osm_id" in r}

        for biz in tqdm(businesses, desc=f"  Checking", leave=False):
            if biz["osm_id"] in existing_ids:
                continue

            listed_site = biz.get("website", "").strip()
            if listed_site:
                # They claim to have a site — verify it still works
                has_website = check_url(listed_site)
            else:
                has_website = False

            time.sleep(0.3)

            results.append({
                "category":   label,
                "name":       biz["name"],
                "phone":      biz["phone"],
                "email":      biz["email"],
                "city":       biz["city"],
                "address":    biz["address"],
                "has_website": has_website,
                "website":    listed_site,
                "osm_id":     biz["osm_id"],
            })
            existing_ids.add(biz["osm_id"])

        done.add(f"{tag_key}={tag_value}")
        progress["done"]    = list(done)
        progress["results"] = results
        save_progress(progress)

        cat_leads = sum(1 for r in results if r["category"] == label and not r["has_website"])
        tqdm.write(f"  Leads in this category: {cat_leads}")
        time.sleep(DELAY)

    # ── Save Excel ────────────────────────────────────────────────────────────
    print("\nSaving results to Excel...")

    leads = [r for r in results if not r.get("has_website")]
    leads_sorted = sorted(leads, key=lambda x: (x["category"], x["name"]))
    write_excel(OUTPUT_LEADS, leads_sorted, "Потенциални клиенти")
    print(f"  Leads      : {OUTPUT_LEADS}  ({len(leads)} businesses)")

    all_sorted = sorted(results, key=lambda x: (x["category"], x["name"]))
    write_excel(OUTPUT_ALL, all_sorted, "Всички резултати")
    print(f"  All        : {OUTPUT_ALL}  ({len(results)} businesses)")

    print(f"\n{'='*60}")
    print(f"  DONE!")
    print(f"  Total found     : {len(results)}")
    print(f"  Have website    : {len(results) - len(leads)}")
    print(f"  NO website      : {len(leads)}  ← your potential clients")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
